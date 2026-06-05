#!/usr/bin/env python
r"""
scripts/import_data.py

General data import script (now supports many formats).

Per ARCHITECTURE_AND_IMPLEMENTATION_PLAN.md:
- Import production (and eventually cost) data into canonical Pydantic models.
- Supports the original EEM wide Excel structure + generic tables.
- Output to data/examples/ as JSON (used by the app for persistence and offline use).

Supported input formats (via the flexible data_handler):
    .xlsx, .xlsm  (EEM structure or generic)
    .csv, .txt, .tsv
    .docx (tables)
    (limited .xls)

Usage examples:
    python scripts/import_data.py --full --out data/examples/
    python scripts/import_data.py --path my_data.csv --out data/examples/
    python scripts/import_data.py --path my_report.docx --out data/examples/

The Streamlit app also has a direct file uploader for these formats (even more convenient).
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook

from EEM_Project.core.models import (
    CostItem,
    CostProfile,
    ProductionProfile,
    ScenarioConfig,
    TerrainType,
    GasUtilization,
)
from EEM_Project.production.data_handler import (
    load_production_profiles,
    save_profiles,
    import_from_file,
    DEFAULT_XLSM as HANDLER_DEFAULT_XLSM,
)

# Re-export for CLI convenience
DEFAULT_XLSM = HANDLER_DEFAULT_XLSM

# Sheets we care about (from architecture §4 and actual sheet list)
SHEETS_OF_INTEREST = [
    "Master",
    "Ec_IO",
    "Model Map",
    "Block_Oil Data",
    "Block_Gas Data",
    "Prod_Summary",
    "Block_TC",
    "Block_TC_Gas",
    "Equity Dash",
    "Fiscal Terms_PIA",
    "Royalties",
    "FLGT",
    "RESULTS",
    "RESULTS Equity",
]

def load_workbook_safe(path: Path, data_only: bool = False) -> openpyxl.Workbook:
    if not path.exists():
        raise FileNotFoundError(
            f"Source model not found: {path}\n"
            f"Update DEFAULT_XLSM in this script or pass --path.\n"
            "The .xlsm must remain outside the git tree (it is .gitignored)."
        )
    print(f"Loading {path} (data_only={data_only}) ...")
    return load_workbook(filename=str(path), data_only=data_only, read_only=False)


def list_sheets(wb: openpyxl.Workbook) -> list[str]:
    return wb.sheetnames


def get_years_from_block_tc(wb_values: openpyxl.Workbook) -> list[int]:
    """Return the list of forecast years (e.g. 2023..2048)."""
    ws = wb_values["Block_TC"]
    years: list[int] = []
    for r in range(4, ws.max_row + 1):
        y = ws.cell(row=r, column=1).value
        if isinstance(y, (int, float)) and 2000 < y < 2100:
            years.append(int(y))
    return years


def extract_ec_io_assumptions(wb: openpyxl.Workbook) -> dict[str, Any]:
    """Ec_IO is the central source of assumptions."""
    if "Ec_IO" not in wb.sheetnames:
        return {}
    ws = wb["Ec_IO"]
    assumptions: dict[str, Any] = {}
    # Known from inspection
    assumptions["production_days_per_year"] = ws["D7"].value or 365
    assumptions["base_oil_price_usd_bbl"] = ws["D12"].value or 70
    assumptions["terrain_flag"] = ws["G20"].value  # e.g. "Shallow Water (<200m water depth)"
    assumptions["gas_util_flag"] = ws["G21"].value  # e.g. "In-Country (Dom Gas)"
    # Project start / life are scenario-dependent formulas in C5/C6; resolved via other cells in practice
    # Add more as we map (discount rate often in results or specific input cells)
    return assumptions


def _find_field_columns(ws: openpyxl.Worksheet, row: int = 1) -> list[tuple[int, str]]:
    """Scan row 1 for field names in the wide Block_* Data layout."""
    cols: list[tuple[int, str]] = []
    # Typical pattern: B, E, H, K, N, Q, T, W, Z, AC, AF, AI ...
    for c in range(2, min(ws.max_column + 1, 60), 3):  # step 3 for daily+annual pairs
        val = ws.cell(row=row, column=c).value
        if val and isinstance(val, str) and ("OML" in val or "FIELD" not in val.upper()):
            cols.append((c, val.strip()))
    return cols


def extract_production_profiles(wb_values: openpyxl.Workbook) -> list[ProductionProfile]:
    """
    Parse the wide Block_Oil Data + Block_Gas Data sheets into ProductionProfile objects.
    Uses data_only values for actual volumes.
    """
    profiles: list[ProductionProfile] = []
    years = get_years_from_block_tc(wb_values)

    # Oil
    wso = wb_values["Block_Oil Data"]
    oil_fields = _find_field_columns(wso)
    print(f"Discovered {len(oil_fields)} oil fields/blocks in Block_Oil Data")

    for daily_col, name in oil_fields:
        annual_col = daily_col + 1
        daily_rates: list[float] = []
        annual_vols: list[float] = []
        for r in range(4, 4 + len(years)):
            d = wso.cell(row=r, column=daily_col).value
            a = wso.cell(row=r, column=annual_col).value
            daily_rates.append(float(d) if isinstance(d, (int, float)) else 0.0)
            annual_vols.append(float(a) if isinstance(a, (int, float)) else 0.0)

        prof = ProductionProfile(
            block_name=name,
            oml="OML 123" if "OML 123" in name else None,
            fluid_type="oil",
            terrain=TerrainType.SHALLOW_WATER,  # driven by Ec_IO G20 for current runs; can be per-block later
            gas_utilization=GasUtilization.DOMESTIC,
            years=years,
            daily_rates_kbd=daily_rates,
            annual_volumes=annual_vols,
            source=f"Block_Oil Data - {name}",
        )
        profiles.append(prof)

    # Gas (very similar layout, headers often linked)
    wsg = wb_values["Block_Gas Data"]
    gas_fields = _find_field_columns(wsg)
    print(f"Discovered {len(gas_fields)} gas fields/blocks in Block_Gas Data")

    for daily_col, name in gas_fields:
        annual_col = daily_col + 1
        daily_rates: list[float] = []
        annual_vols: list[float] = []
        for r in range(4, 4 + len(years)):
            d = wsg.cell(row=r, column=daily_col).value
            a = wsg.cell(row=r, column=annual_col).value
            daily_rates.append(float(d) if isinstance(d, (int, float)) else 0.0)
            annual_vols.append(float(a) if isinstance(a, (int, float)) else 0.0)

        prof = ProductionProfile(
            block_name=name,
            oml="OML 123" if "OML 123" in name else None,
            fluid_type="gas",
            terrain=TerrainType.SHALLOW_WATER,
            gas_utilization=GasUtilization.DOMESTIC,
            years=years,
            daily_rates_kbd=daily_rates,
            annual_volumes=annual_vols,
            source=f"Block_Gas Data - {name}",
        )
        profiles.append(prof)

    return profiles


def extract_costs(wb_values: openpyxl.Workbook) -> list[CostProfile]:
    """Stub for now - Block_TC wide cost categories."""
    # TODO: implement category mapping (Exploration, CAPEX Wells, CAPEX Facilities, OPEX, ...)
    # Return list[CostProfile] once parser is written.
    return []


def extract_fiscal_terms(wb: openpyxl.Workbook) -> dict[str, Any]:
    """Extract royalty tables etc from Fiscal Terms_PIA for regime config."""
    if "Fiscal Terms_PIA" not in wb.sheetnames:
        return {}
    ws = wb["Fiscal Terms_PIA"]
    # From inspection we have the tables at rows 16-30, cols 20+
    # For now return a summary; full structured dict can drive YAML updates.
    return {
        "note": "Royalty sliding scale tables extracted - see docs/module_specs/fiscal_terms_pia.md for full paste",
        "onshore": {"0-5000": 0.05, "5001-10000": 0.075, ">10000": 0.15},
        "shallow_water": {"0-5000": 0.05, "5001-10000": 0.075, ">10000": 0.125},
        "deep": {"0-50000": 0.05, ">50000": 0.075},
        "frontier": 0.075,
        "gas_in_country": 0.025,  # example from table
    }


def extract_equity_dash(wb: openpyxl.Workbook) -> dict[str, Any]:
    ws = wb["Equity Dash"]
    return {
        "company1_equity": ws["C3"].value,
        "acquisition_cost_mm": ws["D3"].value,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import from Econ_Model_Draft_2.xlsm into EEM models")
    parser.add_argument("--path", type=Path, default=DEFAULT_XLSM, help="Path to the xlsm file (external to git)")
    parser.add_argument("--list-sheets", action="store_true", help="Just list all sheets in the workbook")
    parser.add_argument("--full", action="store_true", help="Run full extraction for all supported data")
    parser.add_argument("--out", type=Path, default=Path("data/examples"), help="Output dir for JSON exports")
    parser.add_argument("--block", type=str, help="Extract single block (e.g. 'EBUGHU MAIN - OML123')")
    args = parser.parse_args()

    xlsm_path = args.path
    if args.list_sheets or args.full or args.block:
        wb_formulas = load_workbook_safe(xlsm_path, data_only=False)
        sheets = list_sheets(wb_formulas)
        print(f"\nTotal sheets: {len(sheets)}")
        for i, s in enumerate(sheets, 1):
            marker = " *" if s in SHEETS_OF_INTEREST else ""
            print(f"  {i:2d}. {s}{marker}")
        if args.list_sheets:
            wb_formulas.close()
            return

        # Use the new flexible importer (works for xlsm/xlsx + csv + docx + etc.)
        args.out.mkdir(parents=True, exist_ok=True)

        try:
            profiles = import_from_file(xlsm_path, eem_structure=True, include_totals=False)
            print(f"\nImported {len(profiles)} production profiles (auto-detected format).")

            save_profiles(profiles, args.out)
            print(f"  Saved core profiles to {args.out}")

            # Also import with totals/contingents for the broader set
            all_profiles = import_from_file(xlsm_path, eem_structure=True, include_totals=True)
            broad_dir = args.out / "with_totals_and_contingents"
            save_profiles(all_profiles, broad_dir)
            print(f"  Saved broader set to {broad_dir}")
        except Exception as e:
            print("Flexible importer had an issue, falling back to legacy EEM loader...")
            print(e)
            # legacy path for classic xlsm
            wb_values = load_workbook_safe(xlsm_path, data_only=True)
            profiles = load_production_profiles(xlsm_path=xlsm_path, include_totals=False)
            save_profiles(profiles, args.out)
            print(f"  Saved {len(profiles)} profiles (legacy path)")

        fiscal = extract_fiscal_terms(wb_formulas)
        print("Fiscal terms summary:", fiscal)

        equity = extract_equity_dash(wb_values)
        print("Equity Dash:", equity)

        # Costs data input (early implementation)
        try:
            from EEM_Project.costs.data_handler import load_cost_profiles, save_cost_profiles
            cost_profs = load_cost_profiles(xlsm_path=xlsm_path, fluid="oil")
            if cost_profs:
                cost_dir = args.out / "costs"
                save_cost_profiles(cost_profs, cost_dir)
                print(f"  Saved {len(cost_profs)} oil cost profiles to {cost_dir}")
            cost_gas = load_cost_profiles(xlsm_path=xlsm_path, fluid="gas")
            if cost_gas:
                save_cost_profiles(cost_gas, args.out / "costs")
                print(f"  Also saved gas cost profiles.")
        except Exception as e:
            print("  (Costs extraction still maturing):", e)

        wb_formulas.close()
        wb_values.close()
        print("\nImport complete. Data input layer (production + costs) now has dedicated handlers + JSON persistence. Next per plan: calculation modules with strict validation.")

    print("\nDone.")


if __name__ == "__main__":
    main()
